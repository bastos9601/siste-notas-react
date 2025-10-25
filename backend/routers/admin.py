from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from typing import List
from core.database import get_db
from models import Usuario, Alumno, Docente, Asignatura, Nota, matriculas, HistorialAcademico, AsignaturaHistorial, NotaHistorial, ReporteDocente, ReporteArchivoDocente
from schemas import (
    AlumnoCreate, AlumnoUpdate, Alumno as AlumnoSchema,
    DocenteCreate, Docente as DocenteSchema,
    AsignaturaCreate, Asignatura as AsignaturaSchema,
    MatriculaCreate, Matricula,
    ReporteDocente as ReporteDocenteSchema
)
from core.auth import require_role, get_password_hash, verify_password
from fastapi import BackgroundTasks
from sqlalchemy.orm import Session
from starlette.responses import FileResponse
from fastapi.responses import StreamingResponse
import os
import re
import csv
import io
import unicodedata
from sqlalchemy import func

# Reutilizaremos la misma regla de nota mínima que en el router de alumno
PASSING_GRADE = int(os.getenv("PASSING_GRADE", "11"))

router = APIRouter()


def get_next_cycle(ciclo: str) -> str:
    text = str(ciclo).strip()
    m = re.search(r"(\d+)(?!.*\d)", text)
    if m:
        current = int(m.group(1))
        next_cycle = current + 1
        next_ciclo_str = re.sub(r"(\d+)(?!.*\d)", str(next_cycle), text, count=1)
        return next_ciclo_str

    roman_map = {
        'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5,
        'VI': 6, 'VII': 7, 'VIII': 8, 'IX': 9, 'X': 10
    }
    m2 = re.search(r"\b(I|II|III|IV|V|VI|VII|VIII|IX|X)\b", text, flags=re.IGNORECASE)
    if m2:
        roman = m2.group(1).upper()
        current = roman_map.get(roman)
        if not current:
            raise ValueError("No se pudo determinar el siguiente ciclo (roman not supported)")
        next_num = current + 1
        inv_map = {v: k for k, v in roman_map.items()}
        next_roman = inv_map.get(next_num, str(next_num))
        next_ciclo_str = re.sub(r"\b(I|II|III|IV|V|VI|VII|VIII|IX|X)\b", next_roman, text, count=1, flags=re.IGNORECASE)
        return next_ciclo_str

    raise ValueError("No se pudo determinar el siguiente ciclo a partir del valor de 'ciclo'.")

# Nuevo helper: obtener ciclo base sin sección
def get_base_ciclo(ciclo: str) -> str:
    text = str(ciclo).strip()
    m2 = re.search(r"\b(I|II|III|IV|V|VI|VII|VIII|IX|X)\b", text, flags=re.IGNORECASE)
    if m2:
        return m2.group(1).upper()
    m = re.search(r"(\d+)(?!.*\d)", text)
    if m:
        return m.group(1)
    return text


def alumno_aprobo_asignatura(db: Session, alumno_id: int, asignatura_id: int) -> bool:
    notas = db.query(Nota).filter(
        Nota.alumno_id == alumno_id,
        Nota.asignatura_id == asignatura_id,
        Nota.publicada == True
    ).all()
    if not notas:
        return False
    return max(n.calificacion for n in notas) >= PASSING_GRADE


def registrar_alumno_en_siguiente_ciclo(db: Session, alumno: Alumno) -> dict:
    resultado = {"alumno_id": alumno.id, "nombre": alumno.nombre_completo, "matriculado": False, "registrado": False, "mensaje": ""}
    try:
        next_ciclo = get_next_cycle(alumno.ciclo)
    except ValueError as e:
        resultado["mensaje"] = str(e)
        return resultado

    # Guardar el ciclo actual antes de cambiarlo
    ciclo_actual = alumno.ciclo

    matriculas_data = db.execute(
        matriculas.select().where(matriculas.c.alumno_id == alumno.id)
    ).fetchall()

    asignaturas_actuales_ids = [m.asignatura_id for m in matriculas_data]
    asignaturas_actuales = db.query(Asignatura).filter(
        Asignatura.id.in_(asignaturas_actuales_ids),
        Asignatura.ciclo == get_base_ciclo(ciclo_actual)
    ).all()

    if not asignaturas_actuales:
        resultado["mensaje"] = "No se encontraron asignaturas del ciclo actual para evaluar."
        return resultado

    # Verificar si todas las asignaturas están aprobadas
    asignaturas_no_aprobadas = []
    for asign in asignaturas_actuales:
        if not alumno_aprobo_asignatura(db, alumno.id, asign.id):
            asignaturas_no_aprobadas.append(asign.nombre)
    
    if asignaturas_no_aprobadas:
        resultado["mensaje"] = f"No puede avanzar al siguiente ciclo. No ha aprobado las siguientes asignaturas: {', '.join(asignaturas_no_aprobadas)}."
        resultado["puede_avanzar"] = False
        return resultado
    
    resultado["puede_avanzar"] = True

    # Crear historial académico para el ciclo actual antes de avanzar al siguiente
    historial = HistorialAcademico(
        alumno_id=alumno.id,
        ciclo=ciclo_actual
    )
    db.add(historial)
    db.flush()  # Para obtener el ID del historial
    
    # Agregar asignaturas al historial
    for asignatura in asignaturas_actuales:
        # Calcular promedio de notas para esta asignatura
        promedio_query = db.query(func.avg(Nota.calificacion)).filter(
            Nota.alumno_id == alumno.id,
            Nota.asignatura_id == asignatura.id
        ).scalar()
        
        promedio = promedio_query if promedio_query else 0.0
        
        # Crear asignatura en historial
        asignatura_historial = AsignaturaHistorial(
            historial_id=historial.id,
            nombre=asignatura.nombre,
            promedio=promedio
        )
        db.add(asignatura_historial)
        db.flush()
        
        # Obtener notas de esta asignatura
        notas = db.query(Nota).filter(
            Nota.alumno_id == alumno.id,
            Nota.asignatura_id == asignatura.id
        ).all()
        
        # Guardar notas en historial
        for nota in notas:
            nota_historial = NotaHistorial(
                asignatura_id=asignatura_historial.id,
                calificacion=nota.calificacion,
                tipo_nota=nota.tipo_nota,
                fecha_registro=nota.fecha_registro
            )
            db.add(nota_historial)
    


    asignaturas_siguiente = db.query(Asignatura).filter(Asignatura.ciclo == get_base_ciclo(next_ciclo)).all()

    # Actualizar solo el campo ciclo del alumno (registrar avance de ciclo)
    alumno.ciclo = next_ciclo
    # Eliminar las matrículas del ciclo anterior para no trasladarlas
    if asignaturas_actuales_ids:
        db.execute(
            matriculas.delete().where(
                matriculas.c.alumno_id == alumno.id,
                matriculas.c.asignatura_id.in_(asignaturas_actuales_ids)
            )
        )
    db.commit()

    asignaturas_siguiente_ids = [a.id for a in asignaturas_siguiente]
    resultado["matriculado"] = False
    resultado["registrado"] = True
    resultado["asignaturas_siguiente_ids"] = asignaturas_siguiente_ids
    if asignaturas_siguiente_ids:
        resultado["mensaje"] = f"Registrado en el siguiente ciclo ({next_ciclo}). Se generó el historial del ciclo anterior. No se trasladaron matrículas ni asignaturas."
    else:
        resultado["mensaje"] = f"Registrado en el siguiente ciclo ({next_ciclo}). Se generó el historial del ciclo anterior. No se trasladaron matrículas ni asignaturas."
    resultado["puede_avanzar"] = True

    return resultado


@router.post("/alumnos/{alumno_id}/registrar-siguiente-ciclo")
async def admin_registrar_siguiente_ciclo_alumno(
    alumno_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Endpoint admin para registrar un alumno en el siguiente ciclo sin matricularlo."""
    alumno = db.query(Alumno).filter(Alumno.id == alumno_id).first()
    if not alumno:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Alumno no encontrado")
    resultado = registrar_alumno_en_siguiente_ciclo(db, alumno)
    return resultado


@router.post("/registrar-siguiente-ciclo/todos")
async def admin_registrar_siguiente_ciclo_todos(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Endpoint admin para registrar todos los alumnos en el siguiente ciclo sin matricularlos.
    Retorna un reporte con los resultados por alumno.
    """
    alumnos = db.query(Alumno).all()
    reporte = []
    for alumno in alumnos:
        reporte.append(registrar_alumno_en_siguiente_ciclo(db, alumno))
    return reporte


# ========== GESTIÓN DE ALUMNOS ==========

@router.post("/alumnos", response_model=AlumnoSchema)
async def crear_alumno(
    alumno_data: AlumnoCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Crear nuevo alumno"""
    print(f"DEBUG: Datos recibidos: {alumno_data}")
    
    # Verificar si el DNI ya existe
    existing_alumno = db.query(Alumno).filter(Alumno.dni == alumno_data.dni).first()
    if existing_alumno:
        print(f"DEBUG: DNI ya existe: {alumno_data.dni}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El DNI ya está registrado"
        )
    
    # Verificar si el email ya existe
    existing_user = db.query(Usuario).filter(Usuario.email == alumno_data.email).first()
    if existing_user:
        print(f"DEBUG: Email ya existe: {alumno_data.email}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El email ya está registrado"
        )
    
    # Crear usuario
    hashed_password = get_password_hash(alumno_data.password)
    db_user = Usuario(
        nombre=alumno_data.nombre_completo.split()[0],  # Primer nombre
        email=alumno_data.email,
        password_hash=hashed_password,
        rol="alumno"
    )
    
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    
    # Crear alumno
    db_alumno = Alumno(
        nombre_completo=alumno_data.nombre_completo,
        dni=alumno_data.dni,
        fecha_nacimiento=alumno_data.fecha_nacimiento,
        genero=alumno_data.genero,
        telefono=alumno_data.telefono,
        ciclo=alumno_data.ciclo,
        usuario_id=db_user.id
    )
    
    db.add(db_alumno)
    db.commit()
    db.refresh(db_alumno)
    
    return db_alumno

@router.put("/alumnos/{alumno_id}", response_model=AlumnoSchema)
async def actualizar_alumno(
    alumno_id: int,
    alumno_data: AlumnoUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Actualizar alumno existente"""
    print(f"DEBUG UPDATE: Actualizando alumno {alumno_id}")
    print(f"DEBUG UPDATE: Datos recibidos: {alumno_data}")
    
    # Buscar el alumno
    db_alumno = db.query(Alumno).filter(Alumno.id == alumno_id).first()
    if not db_alumno:
        print(f"DEBUG UPDATE: Alumno {alumno_id} no encontrado")
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Alumno no encontrado"
        )
    
    # Verificar si el DNI ya existe en otro alumno
    existing_alumno = db.query(Alumno).filter(
        Alumno.dni == alumno_data.dni,
        Alumno.id != alumno_id
    ).first()
    if existing_alumno:
        print(f"DEBUG UPDATE: DNI {alumno_data.dni} ya existe en otro alumno")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El DNI ya está registrado en otro alumno"
        )
    
    # Verificar si el email ya existe en otro usuario
    existing_user = db.query(Usuario).filter(
        Usuario.email == alumno_data.email,
        Usuario.id != db_alumno.usuario_id
    ).first()
    if existing_user:
        print(f"DEBUG UPDATE: Email {alumno_data.email} ya existe en otro usuario")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El email ya está registrado en otro usuario"
        )
    
    # Actualizar datos del alumno
    db_alumno.nombre_completo = alumno_data.nombre_completo
    db_alumno.dni = alumno_data.dni
    db_alumno.fecha_nacimiento = alumno_data.fecha_nacimiento
    db_alumno.genero = alumno_data.genero
    db_alumno.telefono = alumno_data.telefono
    db_alumno.ciclo = alumno_data.ciclo
    
    # Actualizar datos del usuario
    db_alumno.usuario.nombre = alumno_data.nombre_completo.split()[0]  # Primer nombre
    db_alumno.usuario.email = alumno_data.email
    
    # Actualizar contraseña solo si se proporciona
    if alumno_data.password is not None and alumno_data.password.strip() != "":
        db_alumno.usuario.password_hash = get_password_hash(alumno_data.password)
    
    db.commit()
    db.refresh(db_alumno)
    
    return db_alumno

@router.delete("/alumnos/{alumno_id}")
async def eliminar_alumno(
    alumno_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Eliminar alumno completamente, incluyendo su historial académico"""
    # Buscar el alumno
    db_alumno = db.query(Alumno).filter(Alumno.id == alumno_id).first()
    if not db_alumno:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Alumno no encontrado"
        )
    
    try:
        # Eliminar historial académico completo
        # 1. Primero eliminar las notas del historial
        historiales = db.query(HistorialAcademico).filter(HistorialAcademico.alumno_id == alumno_id).all()
        for historial in historiales:
            # Obtener todas las asignaturas del historial
            asignaturas_historial = db.query(AsignaturaHistorial).filter(AsignaturaHistorial.historial_id == historial.id).all()
            for asignatura in asignaturas_historial:
                # Eliminar todas las notas de la asignatura en el historial
                db.query(NotaHistorial).filter(NotaHistorial.asignatura_id == asignatura.id).delete()
            
            # Eliminar todas las asignaturas del historial
            db.query(AsignaturaHistorial).filter(AsignaturaHistorial.historial_id == historial.id).delete()
        
        # Eliminar todos los registros de historial académico
        db.query(HistorialAcademico).filter(HistorialAcademico.alumno_id == alumno_id).delete()
        
        # Eliminar todas las notas del alumno
        db.query(Nota).filter(Nota.alumno_id == alumno_id).delete()

        # Los promedios se calculan dinámicamente; no existen registros persistidos que eliminar
        
        # Eliminar todas las matrículas del alumno
        db.execute(
            matriculas.delete().where(matriculas.c.alumno_id == alumno_id)
        )
        
        # Guardar el ID del usuario para eliminarlo después
        usuario_id = db_alumno.usuario_id
        
        # Eliminar el alumno
        db.delete(db_alumno)
        db.commit()
        
        # Eliminar el usuario asociado
        db_usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()
        if db_usuario:
            db.delete(db_usuario)
            db.commit()
        
        return {"message": "Alumno eliminado completamente junto con todo su historial académico, notas y matrículas"}
    
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al eliminar el alumno: {str(e)}"
        )

@router.get("/alumnos")
async def listar_alumnos(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Listar todos los alumnos"""
    try:
        alumnos = db.query(Alumno).order_by(Alumno.nombre_completo.asc()).all()
        print(f"DEBUG: Encontrados {len(alumnos)} alumnos")
        
        # Convertir manualmente a diccionario para evitar problemas de serialización
        result = []
        for alumno in alumnos:
            usuario = db.query(Usuario).filter(Usuario.id == alumno.usuario_id).first()
            result.append({
                "id": alumno.id,
                "nombre_completo": alumno.nombre_completo,
                "dni": alumno.dni,
                "fecha_nacimiento": alumno.fecha_nacimiento,
                "genero": alumno.genero,
                "telefono": alumno.telefono,
                "ciclo": alumno.ciclo,
                "usuario_id": alumno.usuario_id,
                "usuario": {
                    "id": usuario.id,
                    "nombre": usuario.nombre,
                    "email": usuario.email,
                    "rol": usuario.rol,
                    "activo": usuario.activo,
                    "fecha_creacion": usuario.fecha_creacion
                } if usuario else None
            })
        
        return result
    except Exception as e:
        print(f"DEBUG ERROR: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener alumnos: {str(e)}"
        )

@router.post("/alumnos/import-csv")
async def importar_alumnos_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Importar alumnos desde un archivo CSV.
    Encabezados aceptados (flexibles): nombre_completo | nombre + apellidos, dni | documento,
    ciclo | semestre, email | correo, [password | contraseña], [seccion | sección],
    [fecha_nacimiento], [genero | sexo], [telefono | celular].
    Detecta automáticamente el delimitador (coma, punto y coma, tab)."""
    try:
        raw = await file.read()
        # Intentar decodificar como UTF-8, caer a Latin-1 si falla
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")

        # Detectar delimitador automáticamente
        try:
            sample = text[:4096]
            dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
        except Exception:
            dialect = csv.excel

        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        total = 0
        inserted = 0
        skipped = []
        errors = []

        # Normalizador de encabezados
        def norm(s: str) -> str:
            s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
            s = s.lower().strip()
            for ch in ["_", "-", ".", ";", ":"]:
                s = s.replace(ch, " ")
            s = re.sub(r"\s+", " ", s)
            return s

        fieldnames = reader.fieldnames or []
        norm_to_original = {norm(fn): fn for fn in fieldnames}

        def find_header(*candidates):
            for c in candidates:
                key = norm(c)
                if key in norm_to_original:
                    return norm_to_original[key]
            return None

        for row in reader:
            total += 1

            # Mapear campos flexibles
            h_nombre_completo = find_header(
                "nombre completo", "nombres y apellidos", "nombre y apellidos",
                "alumno", "estudiante", "nombreyapellido"
            )
            h_nombre = find_header("nombre", "nombres")
            h_apellidos = find_header("apellidos", "apellido")
            h_dni = find_header(
                "dni", "documento", "numero documento", "nro documento",
                "num documento", "num doc", "numero de documento", "cedula", "identidad"
            )
            h_ciclo = find_header("ciclo", "semestre", "periodo", "nivel")
            h_seccion = find_header("seccion", "seccion", "seccion", "seccion", "seccion", "seccion", "sección", "grupo", "paralelo", "sec")
            h_email = find_header("email", "correo", "correo electronico", "e mail", "mail")
            h_password = find_header("password", "contrasena", "contraseña", "clave")
            h_telefono = find_header("telefono", "celular", "tel")
            h_genero = find_header("genero", "sexo")
            h_fnac = find_header("fecha nacimiento", "fechanacimiento", "fecha de nacimiento", "nacimiento", "f nacimiento")

            # Obtener valores
            if h_nombre_completo:
                nombre = str(row.get(h_nombre_completo) or "").strip()
            else:
                nombre = (" ".join([
                    str(row.get(h_nombre) or "").strip(),
                    str(row.get(h_apellidos) or "").strip()
                ])).strip()

            dni = str(row.get(h_dni) or "").strip()
            ciclo_raw = str(row.get(h_ciclo) or "").strip()
            email = str(row.get(h_email) or "").strip().lower()
            password = str(row.get(h_password) or "").strip()
            seccion = str(row.get(h_seccion) or "").strip().upper()
            telefono = (row.get(h_telefono) or None)
            genero = (row.get(h_genero) or None)

            # Fecha de nacimiento opcional en diferentes formatos
            fecha_nacimiento = None
            fn_val = row.get(h_fnac) if h_fnac else None
            if fn_val:
                try:
                    from datetime import datetime
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                        try:
                            fecha_nacimiento = datetime.strptime(str(fn_val).strip(), fmt).date()
                            break
                        except Exception:
                            pass
                except Exception:
                    fecha_nacimiento = None

            # Validaciones básicas
            if not nombre or not dni or not ciclo_raw or not email:
                skipped.append({
                    "row": total, "dni": dni, "email": email,
                    "reason": "Campos obligatorios faltantes"
                })
                continue

            # Inferir sección desde el ciclo si no viene explícita
            ciclo_base = get_base_ciclo(ciclo_raw)
            if not seccion:
                msec = re.search(r"([A-Z])$", ciclo_raw.strip(), flags=re.IGNORECASE)
                if msec:
                    seccion = msec.group(1).upper()

            ciclo_final = ciclo_base if not seccion or seccion.lower() in ["", "sin seccion", "sin sección"] else f"{ciclo_base} {seccion.upper()}"

            # Evitar duplicados por DNI o email
            if db.query(Alumno).filter(Alumno.dni == dni).first():
                skipped.append({"row": total, "dni": dni, "email": email, "reason": "DNI ya existente"})
                continue
            if db.query(Usuario).filter(Usuario.email == email).first():
                skipped.append({"row": total, "dni": dni, "email": email, "reason": "Email ya existente"})
                continue

            # Generar password si no viene
            if not password:
                import secrets, string
                chars = string.ascii_letters + string.digits
                password = ''.join(secrets.choice(chars) for _ in range(8))
            hashed_password = get_password_hash(password)

            # Crear usuario y alumno
            db_user = Usuario(
                nombre=nombre.split()[0] if nombre else "Alumno",
                email=email,
                password_hash=hashed_password,
                rol="alumno"
            )
            db.add(db_user)
            db.flush()

            db_alumno = Alumno(
                nombre_completo=nombre,
                dni=dni,
                fecha_nacimiento=fecha_nacimiento,
                genero=genero,
                telefono=telefono,
                ciclo=ciclo_final,
                usuario_id=db_user.id
            )
            db.add(db_alumno)
            db.commit()
            inserted += 1

        return {
            "total_rows": total,
            "inserted": inserted,
            "skipped": len(skipped),
            "skipped_details": skipped,
            "errors": errors
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al importar CSV: {str(e)}"
        )

# Nuevo: Importar alumnos desde Excel (XLS/XLSX)
@router.post("/alumnos/import-excel")
async def importar_alumnos_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Importar alumnos desde un archivo Excel (XLS/XLSX).
    Encabezados aceptados (flexibles): nombre_completo | nombre + apellidos, dni | documento,
    ciclo | semestre, email | correo, [password | contraseña], [seccion | sección],
    [fecha_nacimiento], [genero | sexo], [telefono | celular].
    Usa la primera hoja del libro y la primera fila como encabezados."""
    try:
        raw = await file.read()
        from io import BytesIO
        try:
            from openpyxl import load_workbook
        except Exception:
            # Si openpyxl no está instalado, devolvemos un error claro
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                                detail="Falta dependencia 'openpyxl' para importar Excel. Instálala en backend.")

        wb = load_workbook(filename=BytesIO(raw), data_only=True)
        ws = wb.active

        # Obtener encabezados de la primera fila
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h is not None else "" for h in header_row]

        # Construir filas como diccionarios
        excel_rows = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, h in enumerate(headers):
                row_dict[h] = values[i] if i < len(values) else None
            excel_rows.append(row_dict)

        total = 0
        inserted = 0
        skipped = []
        errors = []

        # Normalizador de encabezados
        def norm(s: str) -> str:
            s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
            s = s.lower().strip()
            for ch in ["_", "-", ".", ";", ":"]:
                s = s.replace(ch, " ")
            s = re.sub(r"\s+", " ", s)
            return s

        fieldnames = headers or []
        norm_to_original = {norm(fn): fn for fn in fieldnames}

        def find_header(*candidates):
            for c in candidates:
                key = norm(c)
                if key in norm_to_original:
                    return norm_to_original[key]
            return None

        for row in excel_rows:
            total += 1

            # Mapear campos flexibles
            h_nombre_completo = find_header(
                "nombre completo", "nombres y apellidos", "nombre y apellidos",
                "alumno", "estudiante", "nombreyapellido"
            )
            h_nombre = find_header("nombre", "nombres")
            h_apellidos = find_header("apellidos", "apellido")
            h_dni = find_header(
                "dni", "documento", "numero documento", "nro documento",
                "num documento", "num doc", "numero de documento", "cedula", "identidad"
            )
            h_ciclo = find_header("ciclo", "semestre", "periodo", "nivel")
            h_seccion = find_header("seccion", "sección", "grupo", "paralelo", "sec")
            h_email = find_header("email", "correo", "correo electronico", "e mail", "mail")
            h_password = find_header("password", "contrasena", "contraseña", "clave")
            h_telefono = find_header("telefono", "celular", "tel")
            h_genero = find_header("genero", "sexo")
            h_fnac = find_header("fecha nacimiento", "fechanacimiento", "fecha de nacimiento", "nacimiento", "f nacimiento")

            # Obtener valores
            if h_nombre_completo:
                nombre = str(row.get(h_nombre_completo) or "").strip()
            else:
                nombre = (" ".join([
                    str(row.get(h_nombre) or "").strip(),
                    str(row.get(h_apellidos) or "").strip()
                ])).strip()

            dni = str(row.get(h_dni) or "").strip()
            ciclo_raw = str(row.get(h_ciclo) or "").strip()
            email = str(row.get(h_email) or "").strip().lower()
            password = str(row.get(h_password) or "").strip()
            seccion = str(row.get(h_seccion) or "").strip().upper()
            telefono = (row.get(h_telefono) or None)
            genero = (row.get(h_genero) or None)

            # Fecha de nacimiento opcional
            fecha_nacimiento = None
            fn_val = row.get(h_fnac) if h_fnac else None
            if fn_val:
                try:
                    from datetime import datetime
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
                        try:
                            fecha_nacimiento = datetime.strptime(str(fn_val).strip(), fmt).date()
                            break
                        except Exception:
                            pass
                except Exception:
                    fecha_nacimiento = None

            # Validaciones básicas
            if not nombre or not dni or not ciclo_raw or not email:
                skipped.append({
                    "row": total, "dni": dni, "email": email,
                    "reason": "Campos obligatorios faltantes"
                })
                continue

            # Inferir sección desde el ciclo si no viene explícita
            ciclo_base = get_base_ciclo(ciclo_raw)
            if not seccion:
                msec = re.search(r"([A-Z])$", ciclo_raw.strip(), flags=re.IGNORECASE)
                if msec:
                    seccion = msec.group(1).upper()

            ciclo_final = ciclo_base if not seccion or seccion.lower() in ["", "sin seccion", "sin sección"] else f"{ciclo_base} {seccion.upper()}"

            # Evitar duplicados por DNI o email
            if db.query(Alumno).filter(Alumno.dni == dni).first():
                skipped.append({"row": total, "dni": dni, "email": email, "reason": "DNI ya existente"})
                continue
            if db.query(Usuario).filter(Usuario.email == email).first():
                skipped.append({"row": total, "dni": dni, "email": email, "reason": "Email ya existente"})
                continue

            # Generar password si no viene
            if not password:
                import secrets, string
                chars = string.ascii_letters + string.digits
                password = ''.join(secrets.choice(chars) for _ in range(8))
            hashed_password = get_password_hash(password)

            # Crear usuario y alumno
            db_user = Usuario(
                nombre=nombre.split()[0] if nombre else "Alumno",
                email=email,
                password_hash=hashed_password,
                rol="alumno"
            )
            db.add(db_user)
            db.flush()

            db_alumno = Alumno(
                nombre_completo=nombre,
                dni=dni,
                fecha_nacimiento=fecha_nacimiento,
                genero=genero,
                telefono=telefono,
                ciclo=ciclo_final,
                usuario_id=db_user.id
            )
            db.add(db_alumno)
            db.commit()
            inserted += 1

        return {
            "total_rows": total,
            "inserted": inserted,
            "skipped": len(skipped),
            "skipped_details": skipped,
            "errors": errors
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al importar Excel: {str(e)}"
        )

@router.get("/alumnos/{alumno_id}", response_model=AlumnoSchema)
async def obtener_alumno(
    alumno_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Obtener alumno por ID"""
    alumno = db.query(Alumno).filter(Alumno.id == alumno_id).first()
    if not alumno:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Alumno no encontrado"
        )
    return alumno

# ========== GESTIÓN DE DOCENTES ==========

@router.post("/docentes", response_model=DocenteSchema)
async def crear_docente(
    docente_data: DocenteCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Crear nuevo docente"""
    # Verificar si el DNI ya existe
    existing_docente = db.query(Docente).filter(Docente.dni == docente_data.dni).first()
    if existing_docente:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El DNI ya está registrado"
        )
    
    # Verificar si el email ya existe
    existing_user = db.query(Usuario).filter(Usuario.email == docente_data.email).first()
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El email ya está registrado"
        )
    
    # Crear usuario
    hashed_password = get_password_hash(docente_data.password)
    db_user = Usuario(
        nombre=docente_data.nombre_completo.split()[0],  # Primer nombre
        email=docente_data.email,
        password_hash=hashed_password,
        rol="docente"
    )
    
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    
    # Crear docente
    db_docente = Docente(
        nombre_completo=docente_data.nombre_completo,
        dni=docente_data.dni,
        usuario_id=db_user.id
    )
    
    db.add(db_docente)
    db.commit()
    db.refresh(db_docente)
    
    return db_docente

@router.get("/docentes", response_model=List[DocenteSchema])
async def listar_docentes(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Listar todos los docentes"""
    docentes = db.query(Docente).all()
    return docentes

@router.get("/docentes/{docente_id}", response_model=DocenteSchema)
async def obtener_docente(
    docente_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Obtener docente por ID"""
    docente = db.query(Docente).filter(Docente.id == docente_id).first()
    if not docente:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Docente no encontrado"
        )
    return docente

@router.put("/docentes/{docente_id}", response_model=DocenteSchema)
async def actualizar_docente(
    docente_id: int,
    docente_data: DocenteCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Actualizar docente"""
    docente = db.query(Docente).filter(Docente.id == docente_id).first()
    if not docente:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Docente no encontrado"
        )
    
    # Verificar si el DNI ya existe en otro docente
    existing_docente = db.query(Docente).filter(
        Docente.dni == docente_data.dni,
        Docente.id != docente_id
    ).first()
    if existing_docente:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El DNI ya está registrado en otro docente"
        )
    
    # Verificar si el email ya existe en otro usuario
    existing_user = db.query(Usuario).filter(
        Usuario.email == docente_data.email,
        Usuario.id != docente.usuario_id
    ).first()
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El email ya está registrado"
        )
    
    # Actualizar usuario asociado
    usuario = db.query(Usuario).filter(Usuario.id == docente.usuario_id).first()
    if usuario:
        usuario.nombre = docente_data.nombre_completo.split()[0]
        usuario.email = docente_data.email
        if docente_data.password:
            usuario.password_hash = get_password_hash(docente_data.password)
    
    # Actualizar docente
    docente.nombre_completo = docente_data.nombre_completo
    docente.dni = docente_data.dni
    
    db.commit()
    db.refresh(docente)
    return docente

@router.delete("/docentes/{docente_id}")
async def eliminar_docente(
    docente_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Eliminar docente"""
    docente = db.query(Docente).filter(Docente.id == docente_id).first()
    if not docente:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Docente no encontrado"
        )
    
    # Verificar si el docente tiene asignaturas asignadas
    asignaturas = db.query(Asignatura).filter(Asignatura.docente_id == docente_id).all()
    if asignaturas:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se puede eliminar el docente porque tiene asignaturas asignadas"
        )
    
    # Eliminar usuario asociado
    usuario = db.query(Usuario).filter(Usuario.id == docente.usuario_id).first()
    if usuario:
        db.delete(usuario)
    
    db.delete(docente)
    db.commit()
    return {"message": "Docente eliminado correctamente"}

# ========== GESTIÓN DE ASIGNATURAS ==========

@router.post("/asignaturas", response_model=dict)
async def crear_asignatura(
    asignatura_data: AsignaturaCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Crear nueva asignatura y matricular automáticamente a los alumnos del ciclo correspondiente"""
    # Verificar que el docente existe
    docente = db.query(Docente).filter(Docente.id == asignatura_data.docente_id).first()
    if not docente:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Docente no encontrado"
        )
    
    # Crear asignatura
    db_asignatura = Asignatura(
        nombre=asignatura_data.nombre,
        ciclo=asignatura_data.ciclo,
        docente_id=asignatura_data.docente_id
    )
    
    db.add(db_asignatura)
    db.commit()
    db.refresh(db_asignatura)
    
    # Recargar con la relación del docente
    db_asignatura = db.query(Asignatura).options(joinedload(Asignatura.docente)).filter(Asignatura.id == db_asignatura.id).first()
    
    # Matricular automáticamente a todos los alumnos del ciclo correspondiente
    alumnos_ciclo = db.query(Alumno).filter(Alumno.ciclo == asignatura_data.ciclo).all()
    matriculas_creadas = []
    
    for alumno in alumnos_ciclo:
        # Verificar si ya está matriculado en esta asignatura
        existing_matricula = db.execute(
            matriculas.select().where(
                matriculas.c.alumno_id == alumno.id,
                matriculas.c.asignatura_id == db_asignatura.id
            )
        ).first()
        
        # Si no está matriculado, crear la matrícula
        if not existing_matricula:
            db.execute(
                matriculas.insert().values(
                    alumno_id=alumno.id,
                    asignatura_id=db_asignatura.id
                )
            )
            matriculas_creadas.append({
                "alumno_id": alumno.id,
                "alumno_nombre": alumno.nombre_completo,
                "asignatura_id": db_asignatura.id
            })
    
    # Confirmar los cambios en la base de datos
    db.commit()
    
    return {
        "asignatura": {
            "id": db_asignatura.id,
            "nombre": db_asignatura.nombre,
            "ciclo": db_asignatura.ciclo,
            "docente_id": db_asignatura.docente_id,
            "docente": {
                "id": docente.id,
                "nombre_completo": docente.nombre_completo,
                "dni": docente.dni,
                "usuario_id": docente.usuario_id
            }
        },
        "matriculas_automaticas": {
            "total_alumnos_matriculados": len(matriculas_creadas),
            "detalle": matriculas_creadas
        },
        "mensaje": f"Asignatura creada y {len(matriculas_creadas)} alumnos matriculados automáticamente"
    }

@router.get("/asignaturas")
async def listar_asignaturas(
    ciclo: str = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Listar todas las asignaturas, opcionalmente filtradas por ciclo"""
    try:
        # Consulta simple
        query = db.query(Asignatura)
        if ciclo:
            query = query.filter(Asignatura.ciclo == get_base_ciclo(ciclo))
        asignaturas = query.all()
        
        # Convertir a diccionario manualmente
        result = []
        for asignatura in asignaturas:
            docente = db.query(Docente).filter(Docente.id == asignatura.docente_id).first()
            result.append({
                "id": asignatura.id,
                "nombre": asignatura.nombre,
                "ciclo": asignatura.ciclo,
                "docente_id": asignatura.docente_id,
                "docente": {
                    "id": docente.id,
                    "nombre_completo": docente.nombre_completo,
                    "dni": docente.dni,
                    "usuario_id": docente.usuario_id
                } if docente else None
            })
        
        return result
    except Exception as e:
        print(f"Error en listar_asignaturas: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/asignaturas/{asignatura_id}", response_model=AsignaturaSchema)
async def obtener_asignatura(
    asignatura_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Obtener asignatura por ID"""
    asignatura = db.query(Asignatura).options(joinedload(Asignatura.docente)).filter(Asignatura.id == asignatura_id).first()
    if not asignatura:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asignatura no encontrada"
        )
    return asignatura

@router.put("/asignaturas/{asignatura_id}", response_model=AsignaturaSchema)
async def actualizar_asignatura(
    asignatura_id: int,
    asignatura_data: AsignaturaCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Actualizar asignatura existente"""
    db_asignatura = db.query(Asignatura).filter(Asignatura.id == asignatura_id).first()
    if not db_asignatura:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asignatura no encontrada"
        )
    
    # Verificar que el docente existe
    docente = db.query(Docente).filter(Docente.id == asignatura_data.docente_id).first()
    if not docente:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Docente no encontrado"
        )
    
    # Actualizar asignatura
    db_asignatura.nombre = asignatura_data.nombre
    db_asignatura.ciclo = asignatura_data.ciclo
    db_asignatura.docente_id = asignatura_data.docente_id
    
    db.commit()
    db.refresh(db_asignatura)
    
    # Recargar con la relación del docente
    db_asignatura = db.query(Asignatura).options(joinedload(Asignatura.docente)).filter(Asignatura.id == asignatura_id).first()
    
    return db_asignatura

@router.delete("/asignaturas/{asignatura_id}")
async def eliminar_asignatura(
    asignatura_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Eliminar asignatura"""
    # Buscar la asignatura
    db_asignatura = db.query(Asignatura).filter(Asignatura.id == asignatura_id).first()
    if not db_asignatura:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asignatura no encontrada"
        )
    
    # Verificar si la asignatura tiene notas registradas
    notas_count = db.query(Nota).filter(Nota.asignatura_id == asignatura_id).count()
    if notas_count > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"No se puede eliminar la asignatura porque tiene {notas_count} nota(s) registrada(s). Elimine las notas primero."
        )
    
    # Verificar si la asignatura tiene matrículas
    matriculas_count = db.execute(
        matriculas.select().where(matriculas.c.asignatura_id == asignatura_id)
    ).fetchall()
    if matriculas_count:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"No se puede eliminar la asignatura porque tiene {len(matriculas_count)} matrícula(s) activa(s). Elimine las matrículas primero."
        )
    
    # Eliminar la asignatura
    db.delete(db_asignatura)
    db.commit()
    
    return {"message": "Asignatura eliminada correctamente"}

# ========== GESTIÓN DE MATRÍCULAS ==========

@router.post("/matriculas", response_model=dict)
async def matricular_alumno(
    matricula_data: MatriculaCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Matricular alumno en todas las asignaturas de su ciclo automáticamente y enviar contraseña temporal."""
    # Verificar que el alumno existe
    alumno = db.query(Alumno).filter(Alumno.id == matricula_data.alumno_id).first()
    if not alumno:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Alumno no encontrado"
        )
    
    # Obtener todas las asignaturas del ciclo del alumno
    asignaturas_ciclo = db.query(Asignatura).filter(Asignatura.ciclo == get_base_ciclo(alumno.ciclo)).all()
    
    if not asignaturas_ciclo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No se encontraron asignaturas para el {alumno.ciclo}"
        )
    
    # Lista para almacenar las matrículas creadas
    matriculas_creadas = []
    
    # Matricular al alumno en todas las asignaturas de su ciclo
    for asignatura in asignaturas_ciclo:
        # Verificar si ya está matriculado en esta asignatura
        existing_matricula = db.execute(
            matriculas.select().where(
                matriculas.c.alumno_id == alumno.id,
                matriculas.c.asignatura_id == asignatura.id
            )
        ).first()
        
        # Si no está matriculado, crear la matrícula
        if not existing_matricula:
            db.execute(
                matriculas.insert().values(
                    alumno_id=alumno.id,
                    asignatura_id=asignatura.id
                )
            )
            matriculas_creadas.append({
                "alumno_id": alumno.id,
                "asignatura_id": asignatura.id,
                "asignatura_nombre": asignatura.nombre
            })
    
    # Confirmar los cambios en la base de datos
    db.commit()
    
    # Obtener el usuario asociado al alumno
    usuario = db.query(Usuario).filter(Usuario.id == alumno.usuario_id).first()
    
    # Información sobre las matrículas creadas
    resultado = {
        "mensaje": f"Alumno matriculado automáticamente en {len(matriculas_creadas)} asignaturas del {alumno.ciclo}",
        "alumno": alumno.nombre_completo,
        "ciclo": alumno.ciclo,
        "matriculas": matriculas_creadas,
        "email_sent": False
    }
    
    # Generar y enviar contraseña temporal
    if usuario:
        # Generar nueva contraseña temporal
        import secrets
        import string
        
        # Generar contraseña temporal de 8 caracteres
        password_chars = string.ascii_letters + string.digits
        temp_password = ''.join(secrets.choice(password_chars) for _ in range(8))
        
        # Actualizar la contraseña en la base de datos
        usuario.password_hash = get_password_hash(temp_password)
        db.commit()
        
        # Intentar enviar email
        try:
            from services.email_service import send_password_email
            email_result = await send_password_email(
                email=usuario.email,
                nombre=alumno.nombre_completo,
                temp_password=temp_password
            )
            
            if email_result["success"]:
                resultado["email_sent"] = True
                resultado["email_message"] = "Contraseña temporal enviada por email"
            else:
                # Si falla el email, retornar la contraseña para mostrar al admin
                resultado["temp_password"] = temp_password
                resultado["email_error"] = email_result["message"]
                resultado["instructions"] = "La contraseña se generó correctamente pero no se pudo enviar por email. Configura las credenciales de email en el archivo .env"
        except Exception as e:
            # Si hay error en el servicio de email, retornar la contraseña
            resultado["temp_password"] = temp_password
            resultado["email_error"] = str(e)
            resultado["instructions"] = "La contraseña se generó correctamente pero no se pudo enviar por email debido a un error."
    
    return resultado

@router.get("/matriculas", response_model=List[Matricula])
async def listar_matriculas(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Listar todas las matrículas"""
    matriculas_data = db.execute(matriculas.select()).fetchall()
    result = []
    
    for matricula in matriculas_data:
        alumno = db.query(Alumno).filter(Alumno.id == matricula.alumno_id).first()
        asignatura = db.query(Asignatura).filter(Asignatura.id == matricula.asignatura_id).first()
        result.append(Matricula(
            alumno_id=matricula.alumno_id,
            asignatura_id=matricula.asignatura_id,
            alumno=alumno,
            asignatura=asignatura
        ))
    
    return result

@router.delete("/matriculas/{alumno_id}/{asignatura_id}")
async def eliminar_matricula(
    alumno_id: int,
    asignatura_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Eliminar matrícula de alumno"""
    # Verificar que la matrícula existe
    existing_matricula = db.execute(
        matriculas.select().where(
            matriculas.c.alumno_id == alumno_id,
            matriculas.c.asignatura_id == asignatura_id
        )
    ).first()
    
    if not existing_matricula:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Matrícula no encontrada"
        )
    
    # Eliminar matrícula
    db.execute(
        matriculas.delete().where(
            matriculas.c.alumno_id == alumno_id,
            matriculas.c.asignatura_id == asignatura_id
        )
    )
    db.commit()
    
    return {"message": "Matrícula eliminada correctamente"}

@router.get("/notas")
async def listar_todas_notas(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Listar todas las notas del sistema (solo admin)"""
    # Filtrar notas para que solo se muestren las del ciclo actual de cada alumno
    notas = db.query(Nota).join(Alumno).join(Asignatura).filter(
        Nota.publicada == True
    ).all()
    # Filtrar en Python para igualar ciclo base del alumno
    notas = [n for n in notas if get_base_ciclo(n.alumno.ciclo) == n.asignatura.ciclo]
    
    notas_data = []
    for nota in notas:
        notas_data.append({
            "id": nota.id,
            "alumno_id": nota.alumno_id,
            "alumno_nombre": nota.alumno.nombre_completo,
            "alumno_dni": nota.alumno.dni,
            "alumno_ciclo": nota.alumno.ciclo,
            "asignatura_id": nota.asignatura_id,
            "asignatura_nombre": nota.asignatura.nombre,
            "docente_nombre": nota.asignatura.docente.nombre_completo,
            "calificacion": nota.calificacion,
            "tipo_nota": nota.tipo_nota,
            "fecha_registro": nota.fecha_registro
        })
    
    return notas_data

# ========== DASHBOARD ==========

@router.get("/dashboard")
async def dashboard_admin(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Dashboard del administrador con estadísticas"""
    total_alumnos = db.query(Alumno).count()
    total_docentes = db.query(Docente).count()
    total_asignaturas = db.query(Asignatura).count()
    total_notas = db.query(Nota).count()
    
    return {
        "total_alumnos": total_alumnos,
        "total_docentes": total_docentes,
        "total_asignaturas": total_asignaturas,
        "total_notas": total_notas
    }

@router.post("/alumnos/{alumno_id}/enviar-contrasena")
async def enviar_contrasena_alumno(
    alumno_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Enviar contraseña o enlace de restablecimiento al alumno"""
    # Buscar el alumno
    alumno = db.query(Alumno).filter(Alumno.id == alumno_id).first()
    if not alumno:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Alumno no encontrado"
        )
    
    # Obtener el usuario asociado
    usuario = db.query(Usuario).filter(Usuario.id == alumno.usuario_id).first()
    if not usuario:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Usuario asociado no encontrado"
        )
    
    # Generar nueva contraseña temporal
    import secrets
    import string
    
    # Generar contraseña temporal de 8 caracteres
    password_chars = string.ascii_letters + string.digits
    temp_password = ''.join(secrets.choice(password_chars) for _ in range(8))
    
    # Actualizar la contraseña en la base de datos
    usuario.password_hash = get_password_hash(temp_password)
    db.commit()
    
    # Intentar enviar email
    try:
        from services.email_service import send_password_email
        email_result = await send_password_email(
            email=usuario.email,
            nombre=alumno.nombre_completo,
            temp_password=temp_password
        )
        
        if email_result["success"]:
            return {
                "message": "Contraseña temporal generada y enviada por email",
                "alumno": {
                    "id": alumno.id,
                    "nombre_completo": alumno.nombre_completo,
                    "email": usuario.email
                },
                "email_sent": True,
                "email_message": email_result["message"]
            }
        else:
            # Si falla el email, retornar la contraseña para mostrar al admin
            return {
                "message": "Contraseña temporal generada, pero falló el envío por email",
                "alumno": {
                    "id": alumno.id,
                    "nombre_completo": alumno.nombre_completo,
                    "email": usuario.email,
                    "temp_password": temp_password
                },
                "email_sent": False,
                "email_error": email_result["message"],
                "instructions": "La contraseña se generó correctamente pero no se pudo enviar por email. Configura las credenciales de email en el archivo .env"
            }
    except Exception as e:
        # Si hay error en el servicio de email, retornar la contraseña
        return {
            "message": "Contraseña temporal generada, pero falló el envío por email",
            "alumno": {
                "id": alumno.id,
                "nombre_completo": alumno.nombre_completo,
                "email": usuario.email,
                "temp_password": temp_password
            },
            "email_sent": False,
            "email_error": f"Error en el servicio de email: {str(e)}",
            "instructions": "La contraseña se generó correctamente pero no se pudo enviar por email. Verifica la configuración de email."
        }

@router.post("/docentes/{docente_id}/enviar-contrasena")
async def enviar_contrasena_docente(
    docente_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Enviar contraseña o enlace de restablecimiento al docente"""
    # Buscar el docente
    docente = db.query(Docente).filter(Docente.id == docente_id).first()
    if not docente:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Docente no encontrado"
        )
    
    # Obtener el usuario asociado
    usuario = db.query(Usuario).filter(Usuario.id == docente.usuario_id).first()
    if not usuario:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Usuario asociado no encontrado"
        )
    
    # Generar nueva contraseña temporal
    import secrets
    import string
    
    # Generar contraseña temporal de 8 caracteres
    password_chars = string.ascii_letters + string.digits
    temp_password = ''.join(secrets.choice(password_chars) for _ in range(8))
    
    # Actualizar la contraseña en la base de datos
    usuario.password_hash = get_password_hash(temp_password)
    db.commit()
    
    # Intentar enviar email
    try:
        from services.email_service import send_password_email
        email_result = await send_password_email(
            email=usuario.email,
            nombre=docente.nombre_completo,
            temp_password=temp_password
        )
        
        if email_result["success"]:
            return {
                "message": "Contraseña temporal generada y enviada por email",
                "docente": {
                    "id": docente.id,
                    "nombre_completo": docente.nombre_completo,
                    "email": usuario.email
                },
                "email_sent": True,
                "email_message": email_result["message"]
            }
        else:
            # Si falla el email, retornar la contraseña para mostrar al admin
            return {
                "message": "Contraseña temporal generada, pero falló el envío por email",
                "docente": {
                    "id": docente.id,
                    "nombre_completo": docente.nombre_completo,
                    "email": usuario.email,
                    "temp_password": temp_password
                },
                "email_sent": False,
                "email_error": email_result["message"],
                "instructions": "La contraseña se generó correctamente pero no se pudo enviar por email. Configura las credenciales de email en el archivo .env"
            }
    except Exception as e:
        # Si hay error en el servicio de email, retornar la contraseña
        return {
            "message": "Contraseña temporal generada, pero falló el envío por email",
            "docente": {
                "id": docente.id,
                "nombre_completo": docente.nombre_completo,
                "email": usuario.email,
                "temp_password": temp_password
            },
            "email_sent": False,
            "email_error": f"Error en el servicio de email: {str(e)}",
            "instructions": "La contraseña se generó correctamente pero no se pudo enviar por email. Verifica la configuración de email."
        }

@router.put("/mi-perfil")
async def actualizar_mi_perfil_admin(
    perfil_data: dict,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Actualizar el perfil del admin (email y/o contraseña)"""
    # Verificar que se proporcionó la contraseña actual
    password_actual = perfil_data.get("password_actual")
    if not password_actual:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La contraseña actual es requerida"
        )
    
    # Verificar la contraseña actual
    if not verify_password(password_actual, current_user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La contraseña actual es incorrecta"
        )
    
    # Verificar si se quiere cambiar el email
    nuevo_email = perfil_data.get("nuevo_email")
    if nuevo_email:
        # Verificar que el nuevo email no esté en uso por otro usuario
        existing_user = db.query(Usuario).filter(
            Usuario.email == nuevo_email,
            Usuario.id != current_user.id
        ).first()
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El email ya está registrado por otro usuario"
            )
        
        # Actualizar el email
        current_user.email = nuevo_email
        current_user.nombre = nuevo_email.split('@')[0]  # Usar parte antes del @ como nombre
    
    # Verificar si se quiere cambiar la contraseña
    nueva_password = perfil_data.get("nueva_password")
    if nueva_password:
        # Verificar que la nueva contraseña tiene al menos 6 caracteres
        if len(nueva_password) < 6:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La nueva contraseña debe tener al menos 6 caracteres"
            )
        
        # Actualizar la contraseña
        current_user.password_hash = get_password_hash(nueva_password)
    
    # Verificar que al menos se está cambiando algo
    if not nuevo_email and not nueva_password:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Debe proporcionar un nuevo email y/o nueva contraseña"
        )
    
    db.commit()
    db.refresh(current_user)
    
    return {
        "message": "Perfil actualizado exitosamente",
        "admin": {
            "id": current_user.id,
            "email": current_user.email,
            "rol": current_user.rol
        }
    }

# ========== REPORTES DE DOCENTES ==========

@router.get("/reportes", response_model=List[ReporteDocenteSchema])
async def listar_reportes_docentes(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Listar reportes enviados por docentes."""
    reportes = db.query(ReporteDocente).order_by(ReporteDocente.fecha_envio.desc()).all()
    return reportes

@router.get("/reportes/{reporte_id}/archivo")
async def descargar_archivo_reporte(
    reporte_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Descargar el archivo del reporte enviado por un docente.
    Si el archivo fue guardado en BD, se transmite en streaming. Si es un path, se sirve desde disco.
    """
    reporte = db.query(ReporteDocente).filter(ReporteDocente.id == reporte_id).first()
    if not reporte:
        raise HTTPException(status_code=404, detail="Reporte no encontrado")

    # Si el archivo fue guardado en BD
    if isinstance(reporte.archivo_path, str) and reporte.archivo_path.startswith("db:"):
        archivo = db.query(ReporteArchivoDocente).filter(ReporteArchivoDocente.reporte_id == reporte.id).order_by(ReporteArchivoDocente.id.desc()).first()
        if not archivo:
            raise HTTPException(status_code=404, detail="Archivo del reporte no encontrado en la base de datos")
        bytes_io = io.BytesIO(archivo.content)
        media_type = archivo.mime_type or "text/csv"
        headers = {"Content-Disposition": f"attachment; filename=\"{archivo.filename}\""}
        return StreamingResponse(bytes_io, media_type=media_type, headers=headers)

    # Compatibilidad: si es un path en disco
    try:
        return FileResponse(path=reporte.archivo_path, filename=os.path.basename(reporte.archivo_path))
    except Exception:
        raise HTTPException(status_code=500, detail="No se pudo leer el archivo del reporte")

@router.delete("/reportes/{reporte_id}")
async def eliminar_reporte_docente(
    reporte_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role("admin"))
):
    """Eliminar un reporte de docente, removiendo su registro y archivo CSV."""
    reporte = db.query(ReporteDocente).filter(ReporteDocente.id == reporte_id).first()
    if not reporte:
        raise HTTPException(status_code=404, detail="Reporte no encontrado")

    # Eliminar archivo según origen (BD o disco)
    file_removed = False
    try:
        if isinstance(reporte.archivo_path, str) and reporte.archivo_path.startswith("db:"):
            # Eliminar filas de archivo almacenado en BD
            archivos = db.query(ReporteArchivoDocente).filter(ReporteArchivoDocente.reporte_id == reporte.id).all()
            for a in archivos:
                db.delete(a)
            file_removed = True if archivos else False
        else:
            if reporte.archivo_path and os.path.exists(reporte.archivo_path):
                os.remove(reporte.archivo_path)
                file_removed = True
    except Exception:
        # No bloquear la eliminación del registro si falla borrar el archivo
        file_removed = False

    # Eliminar registro en la base de datos
    try:
        db.delete(reporte)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"No se pudo eliminar el reporte: {e}")

    return {
        "message": "Reporte eliminado correctamente",
        "file_removed": file_removed
    }
